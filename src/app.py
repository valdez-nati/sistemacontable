from flask import Flask, render_template, request, redirect, url_for, flash,jsonify,session,make_response, send_file,request
from flask_mysqldb import MySQL
from werkzeug.security import check_password_hash
from config import config
from flask_login import LoginManager, current_user, login_user, logout_user, login_required
from flask_wtf.csrf import CSRFProtect
import pandas as pd
from openpyxl.styles import Alignment

from datetime import datetime
import json
import os
import openpyxl
from xlsxwriter import Workbook
from io import BytesIO  
#Modelos
from models.ModelUser import ModelUser
from models.entities.User import User


from routers.cliente import actualizar, insertar



app = Flask(__name__, static_folder='static', template_folder='templates')


csrf= CSRFProtect(app)
app.secret_key = config
db= MySQL(app)
login_manager_app= LoginManager(app)


@login_manager_app.user_loader
def load_user(id):
    return ModelUser.get_by_id(db,id)

@app.route('/')
def index():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    #consulta = db.query("INSERT INTO operaciones tipo=% ", "login")
    if request.method== 'POST':
        #print(request.form['username'])
        #print(request.form['password'])
        user = User(0, request.form['username'], request.form['contraseña'], idrol=0)
        logeado = ModelUser.login(db, user)
        if logeado != None:
            user.idrol = logeado.idrol 
            insert_audit_log(request.form['username'], 'ÉXITO')
            session['user'] = request.form['username']
            if logeado.contraseña:
                login_user(logeado)
                session['idrol'] = user.idrol 
        
                #consulta = ("UPDATE operaciones set res= %, datosIngresados=% ", "acceso exitoso", user)
                return redirect(url_for('home'))
            else:
                flash("Contraseña invalida...")
                return render_template('auth/login.html')

        else:
            insert_audit_log(request.form['username'], 'FALLIDO')
            flash("Usuario no encontrado...")
            return render_template('auth/login.html')

    else:
        return render_template('auth/login.html')
 
    

def insert_audit_log(username, status):
    mycursor = db.connection.cursor()
    query = "INSERT INTO login_audit_log (user, status) VALUES (%s, %s)"
    values = (username, status)
    
    mycursor.execute(query, values)
    db.connection.commit()
    mycursor.close()    


def status_401(error):
    return redirect(url_for('login')), 401

def status_404(error):
    return "<h1>Pagina no encontrada</h1>", 404

# Pagina principal
@app.route('/home')
@login_required
def home():
    mycursor = db.connection.cursor()
    sql = "SELECT * FROM clientes"
    mycursor.execute(sql)
    data = mycursor.fetchall()
    mycursor.close()
    idrol = session.get('idrol')
    id=idrol
    return render_template('home.html', data=data, id=id )

def filter_clients(query):
    mycursor = db.connection.cursor()
    # You can customize this query based on your requirements
    query = f"SELECT * FROM clientes WHERE nombres LIKE '%{query}%' OR apellidos LIKE '%{query}%' OR razonsocial LIKE '%{query}%' OR correo LIKE '%{query}%' OR ruc LIKE '%{query}%'"
    mycursor.execute(query)
    
    
    filtered_data =  mycursor.fetchall()
    
    return filtered_data

@app.route('/buscarclientes', methods=['GET'])
def buscarclientes():
    query = request.args.get('query')
    filtered_data = filter_clients(query)

    idrol = session.get('idrol')
    id=idrol
   
    return render_template('home.html', data=filtered_data,id=id)


@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return response
    
# @app.errorhandler(401)  
# def unauthorized(error):
#     return redirect(url_for('login'))    

@app.errorhandler(401)  
def unauthorized(e):
    return redirect(url_for('login'), code=307)

    
# Insertar borrar y actualizar clientes
@app.route('/nuevocliente', methods=['GET', 'POST'])
@login_required
def nuevocliente():   
    if request.method == 'POST': #para que inserte en la base de datos cuando trae informacion
        
        idrol = session.get('user')
        user=idrol
        insertar(user)
        flash("Guardado con exito...")
        return redirect(url_for('home'))   


@app.route("/editarcliente",methods= ['POST', 'GET'])# tiene que llamarse igual la funcion y la url/
@login_required
def editarcliente():
    if request.method == 'POST':
        
        idrol = session.get('user')
        user=idrol
        actualizar(user)
        flash("Actualizado con exito...")
        return redirect(url_for('home'))

@app.route("/borrarcliente/<string:ruc>", methods=['GET'])
@login_required
def borrarcliente(ruc):
    
    
        
        flash("Cliente eliminado")
        mycursor = db.connection.cursor() #para asegurar la coneccion y el cierre se usa .connection
        mycursor.execute( "DELETE FROM clientes WHERE ruc= %s",(ruc,))
        db.connection.commit()
        mycursor.close()
        return redirect(url_for('home'))
   
# ---------Carga de los asientos---------------

@app.route('/guardar-idcliente', methods=['GET'])
@login_required
def guardar_idcliente():
    try:
        idcliente = request.args.get('idcliente')
        print("trae el id", idcliente) 

        session['idcliente'] = idcliente

        return 'ok'
    except Exception as e:
        print(f"Error: {e}")
        return 'Error en el servidor', 500  




@app.route("/cargar_diario",methods= ['POST', 'GET'])
@login_required
def cargar_diario():
    try:
        fecha = request.form.get('fecha')
        descripcion = request.form.get('descripcion')
        
        idcliente = session.get('idcliente')
        print("cliente", idcliente)
        mycursor = db.connection.cursor()
        último_año = "SELECT MAX(YEAR(fecha)) FROM asientoregistro WHERE clientefk=%s"
        mycursor.execute(último_año, (idcliente,))
        año_permitido = mycursor.fetchone()[0]
        mycursor.close()
        if request.method == 'POST':
            fecha = request.form.get("fecha")
            session['fecha'] = fecha
            descripcion = request.form.get("descripcion")
            session['descripcion'] = descripcion
            print("desctripcion",descripcion)
            importe = request.form.get("importe")
            descripcion_ultima = ""
            tipo = request.form.get("tipo")
            cuentas = request.form.get("cuentas")
            #para poder sacar el año de la fecha            
            fechaform= datetime.strptime(fecha, "%Y-%m-%d")
            print("Convertio:", fechaform.year)
            añoform= fechaform.year
            print("tipo",tipo)
            asiento = 0  
            print("año formu")
            idrol = session.get('idrol')
            id=idrol
            print("rol",id)
            try:
                mycursor = db.connection.cursor()
                último_asiento = "SELECT numeroasiento, descripcion FROM asientoregistro WHERE clientefk=%s AND YEAR(fecha)=%s ORDER BY idregistro DESC LIMIT 1"
                mycursor.execute(último_asiento, (idcliente, añoform))
                último = mycursor.fetchone()
                
               
            
            except Exception as e:
                print("Error executing the query:", e)
            print("ultimo:",último)
            print("descripcion ultima:",descripcion_ultima)
            print("año permitido",año_permitido)
            if año_permitido is not None  :
                print("entro en el if de año permitido")
                if último is not None: 
                    print("entro en el if de ultimo ")
                    if añoform == año_permitido or id==1 : 
                       
                            descripcion_ultima = último[1]

                            if descripcion != descripcion_ultima:
                                # Incrementar número
                                print("Entro para sumar")
                                mycursor = db.connection.cursor()
                                siguiente_asiento = "SELECT MAX(numeroasiento) + 1 FROM asientoregistro WHERE clientefk=%s AND YEAR(fecha)=%s;" 
                                
                            else: 
                                # Mantener mismo número
                                print("numero de asiento igual")
                                mycursor = db.connection.cursor()
                                siguiente_asiento = "SELECT MAX(numeroasiento) FROM asientoregistro WHERE clientefk=%s AND YEAR(fecha)=%s;"
                                
                            # Ejecutar consulta para obtener número siguiente
                            mycursor.execute(siguiente_asiento, (idcliente, añoform))
                            numeroasiento = mycursor.fetchone()[0]
                         
                    else:
                        # Mensaje de error 
                        flash("No puede insertar para ese año")
                        return redirect(url_for('cargar_diario'))         
                else:
                    print("inserta en periodo")
                    sql = "INSERT INTO periodos (periodo,clientefk) VALUES (%s, %s)"
                    val = (añoform,idcliente)
                    mycursor.execute(sql, val,)
                    db.connection.commit()
                        
                    
                    numeroasiento = 1
            
            else:
                print("si entro en el sino")
              
                
                idasiento = 1
                numeroasiento = 1
            
            if tipo == "debe":
                print("entro en el debe")
                debe = importe
                haber = 0
            else:
                if tipo == "haber":
                    print("entro en el haber")

                    debe = 0
                    haber = importe

            print("Fecha",fecha,"Numero de asiento:",numeroasiento)
            print("Descripcion:",descripcion,"Importe:",importe)
            print("Tipo:",tipo,"Cuenta:",cuentas)
            print("CLIENTE:",idcliente)
          
                            
            print("Debe:",debe,"","Haber:",haber)
            if ( descripcion_ultima is None or descripcion != descripcion_ultima): 
            #antes de insertar obtener el ultimo asientoregistro y insertar en el mayor
                print("entro en el if de descripcion")
                
                sql = "INSERT INTO asientoregistro (fecha,numeroasiento,descripcion,clientefk,periodofk) VALUES (%s,%s, %s, %s, %s)"
                val = (fecha,numeroasiento,descripcion,idcliente,añoform)
                mycursor.execute(sql, val,)
                db.connection.commit()
                
                print("numeroasiento inserttado:",numeroasiento)
                session['numeroasiento'] = numeroasiento
                
                idasiento = mycursor.lastrowid
                print("asiento2:",idasiento)
                session['idasiento'] = idasiento
                
                query = """
                        SELECT idregistro
                        FROM asientoregistro 
                        WHERE clientefk =  %s AND YEAR(fecha)= %s
                        ORDER BY numeroasiento DESC LIMIT 1
                        """
                mycursor.execute(query, (idcliente, añoform, ))
            try:
                mycursor = db.connection.cursor()
                sql_id= "SELECT idregistro FROM asientoregistro WHERE clientefk = %s AND YEAR(fecha) = %s ORDER BY idregistro DESC LIMIT 1;"
                mycursor.execute(sql_id, (idcliente, añoform))
                idregistro= mycursor.fetchone()
            
            except Exception as e:
                print("Error executing the query:", e)
            
            print("Idregistro:",idregistro)
                
                
            if asiento==1  :
                idasiento = session.get('idasiento')
            else:
                if idregistro is not None:
                    idasiento=idregistro[0]
                else  :
                    idasiento=1
            print ("insertado asiento:",idasiento)
            mycursor = db.connection.cursor()
            sql2=  "INSERT INTO asientodetalle (debe,haber,plancuentasfk,asientoregistrofk) VALUES (%s, %s, %s, %s)"
            val2 = (debe,haber,cuentas,idasiento)
            mycursor.execute(sql2, val2,)
            db.connection.commit()
            mycursor.close()
        mycursor = db.connection.cursor()
        sql = "SELECT idplancuenta, codigo, descripcion FROM plancuentas WHERE imputable=1"
        mycursor.execute(sql)
        data = mycursor.fetchall()
        session['data'] = data
                                
        return redirect(url_for('fact',data=data, fecha=fecha, descripcion=descripcion,numeroasiento=numeroasiento))# rediret lleva el nombre de la funcion

     
    except Exception as e:
        # En caso de error, para imprimir o manejar el error de alguna manera
        print(f"Error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})




@app.route("/cargar")
@login_required
def fact():

    
    idcliente = session.get("idcliente")
    id = session.get('idrol')
    list = session.get('list')
    data = session.get('data')
    descripcion = session.get('descripcion')
    fecha = session.get('fecha')
    numeroasiento = session.get('numeroasiento')
    print("CLIENTE DE FACT",idcliente)

    mycursor = db.connection.cursor()
 


    sql = "SELECT idplancuenta, codigo, descripcion FROM plancuentas WHERE imputable=1"
    mycursor.execute(sql)
    data = mycursor.fetchall()
    #print("Data from the database:", data) 
    
    mycursor.close()
    
  
    
 
   
   
    return render_template('cargar_diario.html',data=data, idcliente=idcliente, id=id,fecha=fecha,descripcion=descripcion,numeroasiento=numeroasiento)


#REGISTRO DIARIO

@app.route('/mostrar_registro',methods= ['POST', 'GET'])
@login_required
def mostrar_registro():
    if request.method == 'POST':
        idcliente = session.get('idcliente')
        año = int(request.form.get('year'))   
        print("cliente",idcliente,"año",año)
        session['año'] = año 
        mycursor = db.connection.cursor()
        query = """
           SELECT 
                idregistro,
                fecha,
                numeroasiento,
                descripcion,
                debe,
                haber,
                idetalle,
                idplancuenta
            FROM (
                -- Detalles
                SELECT 
                    ar.idregistro AS idregistro,
                    ar.fecha AS fecha,
                    ar.numeroasiento AS numeroasiento,
                    pc.descripcion AS descripcion,
                    ad.debe AS debe,
                    ad.haber AS haber,
                    ad.idetalle AS idetalle,
                    pc.idplancuenta AS idplancuenta,
                    1 AS Orden
                FROM 
                    asientodetalle ad
                JOIN 
                    asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
                JOIN 
                    plancuentas pc ON ad.plancuentasfk = pc.idplancuenta
                WHERE 
                    ar.clientefk = %s
                    AND YEAR(ar.fecha) = %s

                UNION ALL

                -- Totales
                SELECT 
                    '' AS idregistro,
                    '' AS fecha,
                    '' AS numeroasiento,
                    'Totales' AS descripcion,
                    SUM(debe) AS debe,
                    SUM(haber) AS haber,
                    '' AS idetalle,
                    '' AS idplancuenta,
                    2 AS Orden
                FROM (
                    SELECT 
                        ad.debe,
                        ad.haber
                    FROM 
                        asientodetalle ad
                    JOIN 
                        asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
                    JOIN 
                        plancuentas pc ON ad.plancuentasfk = pc.idplancuenta
                    WHERE 
                        ar.clientefk = %s
                        AND YEAR(ar.fecha) = %s
                ) AS Totales
            ) AS Resultados
            ORDER BY Orden, idregistro;

                """
        mycursor.execute(query, (idcliente, año,idcliente, año,))
        registros = mycursor.fetchall()
        
        print("Consulta de totales", registros)
        session['registros'] = registros
           
        # Consulta para obtener los años únicos asociados al cliente
        mycursor = db.connection.cursor()
        sql_years = "SELECT DISTINCT YEAR(fecha) AS year FROM asientoregistro WHERE clientefk = %s"
        mycursor.execute(sql_years, (idcliente,))
        years_data = mycursor.fetchall()
        
        # Obtén la lista de años desde los resultados
        lis =  [year[0] for year in years_data]
        
        
    
        print("resultado de la vista",registros)
        mycursor.close()

        
        mycursor = db.connection.cursor()
        sql_month = "SELECT DISTINCT fecha FROM asientoregistro WHERE clientefk = %s"
        mycursor.execute(sql_month, (idcliente,))
        month_data = mycursor.fetchall()

        mycursor.close()
        unico = set(me[0].strftime('%Y-%m') for me in month_data)
        lista = list(unico)
        idrol = session.get('idrol')
        id=idrol

   


    #return redirect(url_for('registros', registros=registros))
    return render_template('listar_registro.html', registros=registros, lis=lis,id=id,lista=lista)



@app.route('/registros')
@login_required
def registros():
    
    #idcliente = session.get('idcliente')
    idcliente = session.get('idcliente')
    
    
    # Consulta para obtener los años únicos asociados al cliente
    mycursor = db.connection.cursor()
    sql_years = "SELECT DISTINCT YEAR(fecha) AS year FROM asientoregistro WHERE clientefk = %s"
    mycursor.execute(sql_years, (idcliente,))
    years_data = mycursor.fetchall()
    
    # Obtén la lista de años desde los resultados
    lis =  [year[0] for year in years_data]
    
    
   
    print("resultado de la vista",registros)
    mycursor.close()

    mycursor = db.connection.cursor()
    sql_month = "SELECT DISTINCT fecha FROM asientoregistro WHERE clientefk = %s"
    mycursor.execute(sql_month, (idcliente,))
    month_data = mycursor.fetchall()

    mycursor.close()
    unico = set(me[0].strftime('%Y-%m') for me in month_data)
    lista = list(unico)
    idrol = session.get('idrol')
    id=idrol

    return render_template('listar_registro.html', lis=lis, lista=lista, id=id)

@app.route('/mes_registro',methods= ['POST', 'GET'])
@login_required
def mes_registro():
    if request.method == 'POST':
        idcliente = session.get('idcliente')
        mes = request.form.get('month') 
        session['mes'] = mes 
        print("mes:",mes)
        mycursor = db.connection.cursor()
        fecha = datetime.strptime(mes, "%Y-%m")
        m = str(fecha.strftime("%m"))
        y= fecha.year
        print("m:",m,"y:",y)
        query = """
           SELECT 
                idregistro,
                fecha,
                numeroasiento,
                descripcion,
                debe,
                haber,
                idetalle,
                idplancuenta
            FROM (
                -- Detalles
                SELECT 
                    ar.idregistro AS idregistro,
                    ar.fecha AS fecha,
                    ar.numeroasiento AS numeroasiento,
                    pc.descripcion AS descripcion,
                    ad.debe AS debe,
                    ad.haber AS haber,
                    ad.idetalle AS idetalle,
                    pc.idplancuenta AS idplancuenta,
                    1 AS Orden
                FROM 
                    asientodetalle ad
                JOIN 
                    asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
                JOIN 
                    plancuentas pc ON ad.plancuentasfk = pc.idplancuenta
                WHERE 
                    ar.clientefk = %s
                    AND MONTH(ar.fecha)=%s
                    AND YEAR(ar.fecha) = %s

                UNION ALL

                -- Totales
                SELECT 
                    '' AS idregistro,
                    '' AS fecha,
                    '' AS numeroasiento,
                    'Totales' AS descripcion,
                    SUM(debe) AS debe,
                    SUM(haber) AS haber,
                    '' AS idetalle,
                    '' AS idplancuenta,
                    2 AS Orden
                FROM (
                    SELECT 
                        ad.debe,
                        ad.haber
                    FROM 
                        asientodetalle ad
                    JOIN 
                        asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
                    JOIN 
                        plancuentas pc ON ad.plancuentasfk = pc.idplancuenta
                    WHERE 
                       ar.clientefk = %s
                        AND MONTH(ar.fecha)=%s
                        AND YEAR(ar.fecha) = %s
                ) AS Totales
            ) AS Resultados
            ORDER BY Orden, idregistro;

                """
     
        
        mycursor.execute(query, (idcliente,m, y,idcliente,m, y,))
        registros = mycursor.fetchall()
        if session.get('registros') is None:
            session['registros'] = []

        print("Consulta de totales", registros)
    
        # Consulta para obtener los años únicos asociados al cliente
        mycursor = db.connection.cursor()
        sql_years = "SELECT DISTINCT YEAR(fecha) AS year FROM asientoregistro WHERE clientefk = %s"
        mycursor.execute(sql_years, (idcliente,))
        years_data = mycursor.fetchall()
        
        # Obtén la lista de años desde los resultados
        lis =  [year[0] for year in years_data]
        
        
    
        print("resultado de la vista",registros)
        mycursor.close()

        mycursor = db.connection.cursor()
        sql_month = "SELECT DISTINCT fecha FROM asientoregistro WHERE clientefk = %s"
        mycursor.execute(sql_month, (idcliente,))
        month_data = mycursor.fetchall()

        mycursor.close()
        unico = set(me[0].strftime('%Y-%m') for me in month_data)
        lista = list(unico)
        idrol = session.get('idrol')
        id=idrol

   


    #return redirect(url_for('registros', registros=registros))
    return render_template('listar_registro.html', registros=registros, lis=lis,id=id,lista=lista)



@app.route("/editar_registro",methods= ['POST', 'GET'])# tiene que llamarse igual la funcion y la url/
@login_required
def editar_registro():
    if request.method == 'POST':
        idcliente = session.get('idcliente')
        fecha = request.form.get("fecha")
        descripcion = request.form.get("descripcion")
        debe = request.form.get("debe")
        haber = request.form.get("haber")
        numeroasiento= request.form.get("idregistro")
        registrofk= request.form.get("idregistrofk")
        detalle= request.form.get("idetalle")
        año =  session.get('año')#para que la edicion sea dentro del año seleccionado
        print("registro",registrofk)
        print("detalle",detalle)
        print("AÑO DEL COMBO:",año)
        
        if datetime.strptime(fecha, '%Y-%m-%d').year != año:
          flash("No puedes editar registros en un año diferente al seleccionado", "error")
          return redirect(url_for('registros'))
        else: 
            mycursor = db.connection.cursor()
            sql1 = "UPDATE asientoregistro SET fecha=%s WHERE clientefk=%s AND numeroasiento=%s "
            val1 = (fecha,  idcliente,numeroasiento)
            mycursor.execute(sql1, val1)
            db.connection.commit()
            mycursor.close()
            
            
            mycursor = db.connection.cursor()
            sql2 = "UPDATE asientodetalle SET debe=%s, haber=%s, plancuentasfk=%s WHERE  idetalle=%s AND asientoregistrofk=%s  "
                
            val2 = (debe,  haber,descripcion,detalle,registrofk)
            mycursor.execute(sql2, val2)
            db.connection.commit()
            mycursor.close()
      
       

            return redirect(url_for('registros'))



@app.route("/borrar_asiento/<int:idregistro>", methods=['GET'])
@login_required
def borrar_asiento(idregistro):
    
    # flash("Cliente eliminado")
    mycursor = db.connection.cursor()
    sql_id = """SELECT ad.debe,ad.haber,ad.plancuentasfk,ar.periodofk, ar.clientefk
                FROM asientodetalle ad
                JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
                WHERE ar.idregistro = %s;
                """
    mycursor.execute(sql_id, (idregistro,))
    resultados = mycursor.fetchall()
    for row in resultados:
        debe = row[0]
        haber = row[1]
        plan = row[2]
        periodo = row[3]
        cliente = row[4]
        print("debe:",debe,"haber",haber,"plan:",plan,"año:",periodo,"cliente",cliente)
       
        

    mycursor.execute( "DELETE FROM asientoregistro WHERE idregistro = %s",(idregistro,))
    db.connection.commit()
    mycursor.close()

    idcliente = session.get('idcliente')
    
    mycursor = db.connection.cursor()
    sql_years = "SELECT DISTINCT YEAR(fecha) AS year FROM asientoregistro WHERE clientefk = %s"
    mycursor.execute(sql_years, (idcliente,))
    years_data = mycursor.fetchall()
    
    # Obtén la lista de años desde los resultados
    lis =  [year[0] for year in years_data]
    
    
    
    mycursor.close()

    mycursor = db.connection.cursor()
    sql_month = "SELECT DISTINCT fecha FROM asientoregistro WHERE clientefk = %s"
    mycursor.execute(sql_month, (idcliente,))
    month_data = mycursor.fetchall()

    mycursor.close()
    unico = set(me[0].strftime('%Y-%m') for me in month_data)
    lista = list(unico)
    
    idrol = session.get('idrol')
    id=idrol
    return redirect(url_for('registros', lis=lis, lista=lista, id=id))

@app.route('/exportar_excel', methods=['POST'])
def exportar_excel():
    idcliente = session.get('idcliente')
    año = session.get('año') 
    mes = session.get('mes')  
    fecha = datetime.strptime(mes, "%Y-%m")
    m = str(fecha.strftime("%m"))
    y= fecha.year
    mycursor_cliente = db.connection.cursor()
    sql_cliente = "SELECT nombres, apellidos, razonsocial FROM clientes WHERE idcliente = %s"
    mycursor_cliente.execute(sql_cliente, (idcliente,))
    resultado_cliente = mycursor_cliente.fetchone()
    if resultado_cliente:
        # Verificar si nombres y apellidos están vacíos y usar razonsocial en su lugar
        nombres = resultado_cliente[0] if resultado_cliente[0] else ""
        apellidos = resultado_cliente[1] if resultado_cliente[1] else ""
        razonsocial = resultado_cliente[2] if resultado_cliente[2] else ""

        # Unir nombres y apellidos (si están presentes) o usar razonsocial
        nombre_cliente = f"{nombres} {apellidos}" if nombres or apellidos else razonsocial
    else:
        # Manejar el caso en el que no se encuentre el cliente
        nombre_cliente = "Cliente Desconocido"

    mycursor = db.connection.cursor()
    query = """
     SELECT 
        fecha,
        numeroasiento,
        descripcion,
        debe,
        haber
    FROM (
        -- Detalles
        SELECT 
            ar.fecha AS fecha,
            ar.numeroasiento AS numeroasiento,
            pc.descripcion AS descripcion,
            ad.debe AS debe,
            ad.haber AS haber,
            1 AS Orden
        FROM 
            asientodetalle ad
        JOIN 
            asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
        JOIN 
            plancuentas pc ON ad.plancuentasfk = pc.idplancuenta
        WHERE 
            ar.clientefk = %s
            AND YEAR(ar.fecha) = %s
            AND MONTH(ar.fecha) = %s

        UNION ALL

        -- Totales
        SELECT 
            '' AS fecha,
            '' AS numeroasiento,
            'Totales' AS descripcion,
            SUM(debe) AS debe,
            SUM(haber) AS haber,
            2 AS Orden
        FROM (
            SELECT 
                ad.debe,
                ad.haber
            FROM 
                asientodetalle ad
            JOIN 
                asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            JOIN 
                plancuentas pc ON ad.plancuentasfk = pc.idplancuenta
            WHERE 
                 ar.clientefk = %s
            AND YEAR(ar.fecha) = %s
            AND MONTH(ar.fecha) = %s
        ) AS Totales
    ) AS Resultados
    ORDER BY Orden, fecha, numeroasiento;


    """
   

    mycursor.execute(query, (idcliente, y, m,idcliente, y, m))
    registros = mycursor.fetchall()
    print("registeos:",registros)
    carpeta_destino = 'C:/Users/Naty/OneDrive/Documentos'
    df = pd.DataFrame(registros, columns=['Fecha', 'Asiento', 'Descripcion', 'Debe', 'Haber'])

   # Crear un libro de Excel y una hoja de cálculo
    df = df[['Fecha', 'Asiento', 'Descripcion', 'Debe', 'Haber']]
     # Crear un libro de Excel y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DIARIO"

    # Combinar celdas para el título
    ws.merge_cells('A1:E1')
    ws['A1'] = "DIARIO"
    ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Agregar información del cliente y año
    ws.append([f"Cliente: {nombre_cliente}, Año: {año}"])
    ws.merge_cells('A2:E2')
    ws['A2'].font = openpyxl.styles.Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.append([f"Fecha de Expedición: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.merge_cells('A3:E3')
    ws['A3'].font = openpyxl.styles.Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal="center")

    # Agregar encabezados
    headers = ['Fecha', 'Asiento', 'Descripcion', 'Debe', 'Haber']
    ws.append(headers)
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=7):
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Escribir DataFrame
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Guardar el archivo Excel
    carpeta_destino = 'C:/Users/Naty/OneDrive/Documentos'
    fecha_hora_actual = datetime.now().strftime("%Y%m%d%H%M%S")
    nombre_archivo = f' diario_{nombre_cliente}_{año}_{fecha_hora_actual}.xlsx'
    ruta_archivo = os.path.join(carpeta_destino, nombre_archivo)
    wb.save(ruta_archivo)

    return send_file(ruta_archivo, as_attachment=True)

            


@app.route('/diario_año', methods=['POST'])
def diario_excel():
    idcliente = session.get('idcliente')
    año = session.get('año')  

    mycursor_cliente = db.connection.cursor()
    sql_cliente = "SELECT nombres, apellidos, razonsocial FROM clientes WHERE idcliente = %s"
    mycursor_cliente.execute(sql_cliente, (idcliente,))
    resultado_cliente = mycursor_cliente.fetchone()
    if resultado_cliente:
        # Verificar si nombres y apellidos están vacíos y usar razonsocial en su lugar
        nombres = resultado_cliente[0] if resultado_cliente[0] else ""
        apellidos = resultado_cliente[1] if resultado_cliente[1] else ""
        razonsocial = resultado_cliente[2] if resultado_cliente[2] else ""

        # Unir nombres y apellidos (si están presentes) o usar razonsocial
        nombre_cliente = f"{nombres} {apellidos}" if nombres or apellidos else razonsocial
    else:
        # Manejar el caso en el que no se encuentre el cliente
        nombre_cliente = "Cliente Desconocido"

    mycursor = db.connection.cursor()
    query = """
     SELECT 
        fecha,
        numeroasiento,
        descripcion,
        debe,
        haber
    FROM (
        -- Detalles
        SELECT 
            ar.fecha AS fecha,
            ar.numeroasiento AS numeroasiento,
            pc.descripcion AS descripcion,
            ad.debe AS debe,
            ad.haber AS haber,
            1 AS Orden
        FROM 
            asientodetalle ad
        JOIN 
            asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
        JOIN 
            plancuentas pc ON ad.plancuentasfk = pc.idplancuenta
        WHERE 
            ar.clientefk = %s
            AND YEAR(ar.fecha) = %s

        UNION ALL

        -- Totales
        SELECT 
            '' AS fecha,
            '' AS numeroasiento,
            'Totales' AS descripcion,
            SUM(debe) AS debe,
            SUM(haber) AS haber,
            2 AS Orden
        FROM (
            SELECT 
                ad.debe,
                ad.haber
            FROM 
                asientodetalle ad
            JOIN 
                asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            JOIN 
                plancuentas pc ON ad.plancuentasfk = pc.idplancuenta
            WHERE 
                ar.clientefk =%s
                AND YEAR(ar.fecha) =%s
        ) AS Totales
    ) AS Resultados
    ORDER BY Orden, fecha, numeroasiento;


    """

    mycursor.execute(query, (idcliente, año,idcliente, año))
    registros = mycursor.fetchall()
    print("registeos:",registros)
    carpeta_destino = 'C:/Users/Naty/OneDrive/Documentos'
    df = pd.DataFrame(registros, columns=['Fecha', 'Asiento', 'Descripcion', 'Debe', 'Haber'])


    # Crear un libro de Excel y una hoja de cálculo
    df = df[['Fecha', 'Asiento', 'Descripcion', 'Debe', 'Haber']]
     # Crear un libro de Excel y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DIARIO"

    # Combinar celdas para el título
    ws.merge_cells('A1:E1')
    ws['A1'] = "DIARIO"
    ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Agregar información del cliente y año
    ws.append([f"Cliente: {nombre_cliente}, Año: {año}"])
    ws.merge_cells('A2:E2')
    ws['A2'].font = openpyxl.styles.Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.append([f"Fecha de Expedición: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.merge_cells('A3:E3')
    ws['A3'].font = openpyxl.styles.Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal="center")

    # Agregar encabezados
    headers = ['Fecha', 'Asiento', 'Descripcion', 'Debe', 'Haber']
    ws.append(headers)
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=7):
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Escribir DataFrame
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Guardar el archivo Excel
    carpeta_destino = 'C:/Users/Naty/OneDrive/Documentos'
    fecha_hora_actual = datetime.now().strftime("%Y%m%d%H%M%S")
    nombre_archivo = f' diario_{nombre_cliente}_{año}_{fecha_hora_actual}.xlsx'
    ruta_archivo = os.path.join(carpeta_destino, nombre_archivo)
    wb.save(ruta_archivo)

    return send_file(ruta_archivo, as_attachment=True)

            

#MAYOR

@app.route('/mayor')
def mayor():
    idcliente = session.get('idcliente')
    
    # Consulta para obtener los años únicos asociados al cliente
    mycursor = db.connection.cursor()
    sql_years = "SELECT DISTINCT YEAR(fecha) AS year FROM asientoregistro WHERE clientefk = %s"
    mycursor.execute(sql_years, (idcliente,))
    years_data = mycursor.fetchall()
    
    # Obtén la lista de años desde los resultados
    lis =  [year[0] for year in years_data]
    
    mycursor = db.connection.cursor()
    sql = "SELECT idplancuenta, codigo, descripcion FROM plancuentas WHERE imputable=1"
    mycursor.execute(sql)
    data = mycursor.fetchall()
       
    
    mycursor.close()

    
    idrol = session.get('idrol')
    id=idrol

    return render_template('mayor.html', lis=lis, data=data, id=id)




@app.route('/mes_mayor',methods= ['POST', 'GET'])
def mes_mayor():
    if request.method == 'POST':
        idcliente = session.get('idcliente')
        cuentas = request.form.get("cuentas")
        año = request.form.get("year")
        año = session.get('año')
        mycursor = db.connection.cursor()
        
        query = """
            SELECT ar.fecha AS fecha, 
                pc.descripcion AS cuentas,
                ar.numeroasiento AS numeroasiento, 
                ar.descripcion AS descripcion,
                ad.debe AS debe, ad.haber AS haber, 
                CASE WHEN ad.debe - ad.haber < 0 THEN 0 ELSE ad.debe - ad.haber END AS saldo 
            FROM asientodetalle ad 
            JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro 
            JOIN plancuentas pc ON ad.plancuentasfk = pc.idplancuenta 
            WHERE ar.clientefk = %s AND pc.idplancuenta =%s  AND YEAR(ar.fecha) =%s
            ORDER BY  pc.idplancuenta;
            """

        
        mycursor.execute(query, (idcliente,cuentas, año,))
        libros = mycursor.fetchall()
        print("Consulta de totales", libros)
        

        # Consulta para obtener los años únicos asociados al cliente
        mycursor = db.connection.cursor()
        sql_years = "SELECT DISTINCT YEAR(fecha) AS year FROM asientoregistro WHERE clientefk = %s"
        mycursor.execute(sql_years, (idcliente,))
        years_data = mycursor.fetchall()
        
        # Obtén la lista de años desde los resultados
        lis =  [year[0] for year in years_data]
        
        
        
        mycursor.close()


        
        
        idrol = session.get('idrol')
        id=idrol
        
        mycursor = db.connection.cursor()
        sql = "SELECT idplancuenta, codigo, descripcion FROM plancuentas WHERE imputable=1"
        mycursor.execute(sql)
        data = mycursor.fetchall()
        
        
        mycursor.close()
        response = make_response(render_template('mayor.html', lis=lis, data=data, id=id))
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return render_template('mayor.html',libros=libros,  lis=lis, id=id,data=data)



@app.route('/mayor_excel', methods=['POST'])
def mayor_excel():
    idcliente = session.get('idcliente')
    cuentas = request.form.get("cuentas")
    año = request.form.get("year")
    mycursor_cliente = db.connection.cursor()
    sql_cliente = "SELECT nombres, apellidos, razonsocial FROM clientes WHERE idcliente = %s"
    mycursor_cliente.execute(sql_cliente, (idcliente,))
    resultado_cliente = mycursor_cliente.fetchone()
    if resultado_cliente:
        # Verificar si nombres y apellidos están vacíos y usar razonsocial en su lugar
        nombres = resultado_cliente[0] if resultado_cliente[0] else ""
        apellidos = resultado_cliente[1] if resultado_cliente[1] else ""
        razonsocial = resultado_cliente[2] if resultado_cliente[2] else ""

        # Unir nombres y apellidos (si están presentes) o usar razonsocial
        nombre_cliente = f"{nombres} {apellidos}" if nombres or apellidos else razonsocial
    else:
        # Manejar el caso en el que no se encuentre el cliente
        nombre_cliente = "Cliente Desconocido"

    fecha_hora_actual = datetime.now().strftime("%Y%m%d%H%M%S")
    mycursor = db.connection.cursor()
    query = """
      SELECT ar.fecha AS Fecha, 
 	pc.descripcion AS Cuentas,
    ar.numeroasiento AS Asiento, 
    ar.descripcion AS Descripcion,
    ad.debe AS Debe, 
    ad.haber AS Haber, 
    CASE WHEN ad.debe - ad.haber < 0 THEN 0 ELSE ad.debe - ad.haber END AS Saldo 
    FROM asientodetalle ad JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro JOIN plancuentas pc ON ad.plancuentasfk = 				pc.idplancuenta 
    WHERE ar.clientefk = %s AND pc.idplancuenta =%s  AND YEAR(ar.fecha) =%s
    ORDER BY  pc.idplancuenta;
    """

    mycursor.execute(query, (idcliente,cuentas, año))
    registros = mycursor.fetchall()
    print("registeos:",registros)
    df = pd.DataFrame(registros, columns=['Fecha', 'Cuentas',  'Asiento', 'Descripcion', 'Debe', 'Haber', 'Saldo'])
    # Crear un libro de Excel y una hoja de cálculo
    df = df[['Fecha', 'Cuentas',  'Asiento', 'Descripcion', 'Debe', 'Haber', 'Saldo']]
     # Crear un libro de Excel y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MAYOR"

    # Combinar celdas para el título
    ws.merge_cells('A1:G1')
    ws['A1'] = "MAYOR"
    ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Agregar información del cliente y año
    ws.append([f"Cliente: {nombre_cliente}, Año: {año}"])
    ws.merge_cells('A2:G2')
    ws['A2'].font = openpyxl.styles.Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.append([f"Fecha de Expedición: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.merge_cells('A3:G3')
    ws['A3'].font = openpyxl.styles.Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal="center")

    # Agregar encabezados
    headers = ['Fecha', 'Cuentas', 'Asiento', 'Descripcion', 'Debe', 'Haber', 'Saldo']
    ws.append(headers)
    for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=7):
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Escribir DataFrame
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Guardar el archivo Excel
    carpeta_destino = 'C:/Users/Naty/OneDrive/Documentos'
    fecha_hora_actual = datetime.now().strftime("%Y%m%d%H%M%S")
    nombre_archivo = f'mayor_{nombre_cliente}_{año}_{fecha_hora_actual}.xlsx'
    ruta_archivo = os.path.join(carpeta_destino, nombre_archivo)
    wb.save(ruta_archivo)

    return send_file(ruta_archivo, as_attachment=True)



@app.route('/balance')
def balance():
    
    idcliente = session.get('idcliente')
    
    # Consulta para obtener los años únicos asociados al cliente
    mycursor = db.connection.cursor()
    sql_years = "SELECT DISTINCT YEAR(fecha) AS year FROM asientoregistro WHERE clientefk = %s"
    mycursor.execute(sql_years, (idcliente,))
    years_data = mycursor.fetchall()
    
    # Obtén la lista de años desde los resultados
    lis =  [year[0] for year in years_data]
    print("balance año:",lis)
    mycursor.close()

    mycursor = db.connection.cursor()
    sql_month = "SELECT DISTINCT fecha FROM asientoregistro WHERE clientefk = %s"
    mycursor.execute(sql_month, (idcliente,))
    month_data = mycursor.fetchall()

    mycursor.close()
    unico = set(me[0].strftime('%Y-%m') for me in month_data)
    lista = list(unico)

    print("balance mes:",lista)
    return render_template('listar_balance.html', lis=lis, lista=lista)



@app.route('/balance_año', methods=['POST', 'GET'])
def balance_año():
    if request.method == 'POST':
        idcliente = session.get('idcliente')
        año = str(request.form.get('year'))
        print("año del balance a mostrar", año, "cliente", idcliente)
        session['año'] = año
        try:
            mycursor = db.connection.cursor()
            sql = """
            SELECT 
            pc.codigo AS codigo_cuenta,
            pc.descripcion AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_cuenta
            FROM plancuentas pc  
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE (c.idcliente = %s OR c.idcliente IS NULL) 
            AND (p.periodo = %s OR p.periodo IS NULL)
            GROUP BY pc.codigo, pc.descripcion

            UNION ALL

            SELECT
            '1' AS codigo_cuenta,
            'ACTIVO' AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk  
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) = '1'
            AND (c.idcliente = %s OR c.idcliente IS NULL) 
            AND (p.periodo = %s OR p.periodo IS NULL)  
            GROUP BY codigo_cuenta, descripcion_cuenta

            UNION ALL

            SELECT
            '2' AS codigo_cuenta,
            'PASIVO' AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) = '2'
            AND (c.idcliente = %s OR c.idcliente IS NULL)
            AND (p.periodo = %s OR p.periodo IS NULL) 
            GROUP BY codigo_cuenta, descripcion_cuenta
            UNION ALL


            SELECT
            '3' AS codigo_cuenta, 
            'PATRIMONIO NETO' AS descripcion_cuenta,
            COALESCE(
                SUM(CASE WHEN LEFT(pc.codigo, 1) = '1' THEN ad.debe - ad.haber ELSE 0 END) -  
                SUM(CASE WHEN LEFT(pc.codigo, 1) = '2' THEN ad.debe - ad.haber ELSE 0 END),
            0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk  
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) IN ('1','2') 
            AND (c.idcliente = %s OR c.idcliente IS NULL)
            AND (p.periodo = %s OR p.periodo IS NULL)
            GROUP BY codigo_cuenta, descripcion_cuenta

            UNION ALL

            SELECT 
            '4' AS codigo_cuenta,
            'INGRESOS OPERATIVOS' AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo 
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) = '4'
            AND (c.idcliente = %s OR c.idcliente IS NULL)
            AND (p.periodo = %s OR p.periodo IS NULL)
            GROUP BY codigo_cuenta, descripcion_cuenta  

            UNION ALL

            SELECT
            '5' AS codigo_cuenta,
            'COSTOS OPERATIVOS' AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_total 
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) = '5'
            AND (c.idcliente = %s OR c.idcliente IS NULL) 
            AND (p.periodo = %s OR p.periodo IS NULL)   
            GROUP BY codigo_cuenta, descripcion_cuenta

            UNION ALL 

            SELECT
            '6' AS codigo_cuenta,
            'GANANCIAS (O PÉRDIDAS) BRUTAS EN VENTAS' AS descripcion_cuenta,
            COALESCE(
                SUM(CASE WHEN LEFT(pc.codigo, 1) = '4' THEN ad.debe - ad.haber ELSE 0 END) -
                SUM(CASE WHEN LEFT(pc.codigo, 1) = '5' THEN ad.debe - ad.haber ELSE 0 END), 
            0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro  
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente 
            WHERE LEFT(pc.codigo, 1) IN ('4','5')
            AND (c.idcliente = %s OR c.idcliente IS NULL) 
            AND (p.periodo = %s OR p.periodo IS NULL)  
            GROUP BY codigo_cuenta, descripcion_cuenta

            ORDER BY codigo_cuenta;    """
            
            mycursor.execute(sql, (idcliente,año,idcliente,año,idcliente,año,idcliente,año,idcliente,año,idcliente,año,idcliente,año))
            resultados = mycursor.fetchall()
            print("consulta:", resultados)
           


            # Consulta para obtener los años únicos asociados al cliente
            mycursor = db.connection.cursor()
            sql_years = "SELECT DISTINCT YEAR(fecha) AS year FROM asientoregistro WHERE clientefk = %s"
            mycursor.execute(sql_years, (idcliente,))
            years_data = mycursor.fetchall()

            # Obtén la lista de años desde los resultados
            lis = [year[0] for year in years_data]

            mycursor.close()

            mycursor = db.connection.cursor()
            sql_c = "SELECT idplancuenta, descripcion FROM plancuentas WHERE imputable=1"
            mycursor.execute(sql_c)
            data = mycursor.fetchall()
            mycursor.close()

            return render_template('listar_balance.html', lis=lis, data=data, resultados=resultados)

        except Exception as e:
            print("SQL Error:", str(e))
            return jsonify({'success': False, 'error': str(e)})






@app.route('/procesar_balance',methods= ['POST', 'GET'])
@login_required
def exportar_balance():
    try:
        idcliente = session.get('idcliente')
        mycursor_cliente = db.connection.cursor()
        sql_cliente = "SELECT nombres, apellidos, razonsocial FROM clientes WHERE idcliente = %s"
        mycursor_cliente.execute(sql_cliente, (idcliente,))
        resultado_cliente = mycursor_cliente.fetchone()

        if resultado_cliente:
            nombres, apellidos, razonsocial = resultado_cliente
            nombre_cliente = f"{nombres} {apellidos}" if nombres or apellidos else razonsocial
        else:
            nombre_cliente = "Cliente Desconocido"

        año = str(request.form.get('year'))
        print("año ", año)
        fecha_hora_actual = datetime.now().strftime("%Y%m%d%H%M%S")

        mycursor = db.connection.cursor()
        sql = """
            SELECT 
            pc.codigo AS codigo_cuenta,
            pc.descripcion AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_cuenta
            FROM plancuentas pc  
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE (c.idcliente = %s OR c.idcliente IS NULL) 
            AND (p.periodo = %s OR p.periodo IS NULL)
            GROUP BY pc.codigo, pc.descripcion

            UNION ALL

            SELECT
            '1' AS codigo_cuenta,
            'ACTIVO' AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk  
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) = '1'
            AND (c.idcliente = %s OR c.idcliente IS NULL) 
            AND (p.periodo = %s OR p.periodo IS NULL)  
            GROUP BY codigo_cuenta, descripcion_cuenta

            UNION ALL

            SELECT
            '2' AS codigo_cuenta,
            'PASIVO' AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) = '2'
            AND (c.idcliente = %s OR c.idcliente IS NULL)
            AND (p.periodo = %s OR p.periodo IS NULL) 
            GROUP BY codigo_cuenta, descripcion_cuenta
            UNION ALL


            SELECT
            '3' AS codigo_cuenta, 
            'PATRIMONIO NETO' AS descripcion_cuenta,
            COALESCE(
                SUM(CASE WHEN LEFT(pc.codigo, 1) = '1' THEN ad.debe - ad.haber ELSE 0 END) -  
                SUM(CASE WHEN LEFT(pc.codigo, 1) = '2' THEN ad.debe - ad.haber ELSE 0 END),
            0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk  
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) IN ('1','2') 
            AND (c.idcliente = %s OR c.idcliente IS NULL)
            AND (p.periodo = %s OR p.periodo IS NULL)
            GROUP BY codigo_cuenta, descripcion_cuenta

            UNION ALL

            SELECT 
            '4' AS codigo_cuenta,
            'INGRESOS OPERATIVOS' AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo 
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) = '4'
            AND (c.idcliente = %s OR c.idcliente IS NULL)
            AND (p.periodo = %s OR p.periodo IS NULL)
            GROUP BY codigo_cuenta, descripcion_cuenta  

            UNION ALL

            SELECT
            '5' AS codigo_cuenta,
            'COSTOS OPERATIVOS' AS descripcion_cuenta,
            COALESCE(SUM(ad.debe - ad.haber), 0) AS saldo_total 
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente
            WHERE LEFT(pc.codigo, 1) = '5'
            AND (c.idcliente = %s OR c.idcliente IS NULL) 
            AND (p.periodo = %s OR p.periodo IS NULL)   
            GROUP BY codigo_cuenta, descripcion_cuenta

            UNION ALL 

            SELECT
            '6' AS codigo_cuenta,
            'GANANCIAS (O PÉRDIDAS) BRUTAS EN VENTAS' AS descripcion_cuenta,
            COALESCE(
                SUM(CASE WHEN LEFT(pc.codigo, 1) = '4' THEN ad.debe - ad.haber ELSE 0 END) -
                SUM(CASE WHEN LEFT(pc.codigo, 1) = '5' THEN ad.debe - ad.haber ELSE 0 END), 
            0) AS saldo_total
            FROM plancuentas pc
            LEFT JOIN asientodetalle ad ON pc.idplancuenta = ad.plancuentasfk
            LEFT JOIN asientoregistro ar ON ad.asientoregistrofk = ar.idregistro  
            LEFT JOIN periodos p ON ar.periodofk = p.periodo
            LEFT JOIN clientes c ON ar.clientefk = c.idcliente 
            WHERE LEFT(pc.codigo, 1) IN ('4','5')
            AND (c.idcliente = %s OR c.idcliente IS NULL) 
            AND (p.periodo = %s OR p.periodo IS NULL)  
            GROUP BY codigo_cuenta, descripcion_cuenta

            ORDER BY codigo_cuenta;    """
            
        mycursor.execute(sql, (idcliente,año,idcliente,año,idcliente,año,idcliente,año,idcliente,año,idcliente,año,idcliente,año))
        # Ejecutar y obtener resultados
        resultados = mycursor.fetchall() 
        print(resultados)
        #df = pd.DataFrame(resultados)
        
        
        df = pd.DataFrame(resultados, columns=['Codigo', 'Descripcion', 'Monto'])

        # Reorganizar las columnas según la estructura típica de un balance
        df = df[['Codigo', 'Descripcion', 'Monto']]

       
        # Crear un libro de Excel y una hoja de cálculo
        wb = openpyxl.Workbook()
        ws = wb.active
        # Agregar un título general

        

        # Agregar un título más grande en la segunda fila
        ws['A1'] = "BALANCE GENERAL"
        ws.merge_cells('A1:C1')  # Combina las celdas A2, B2, y C2
        ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)  # Puedes ajustar el tamaño de la fuente según tu preferencia

                # Agregar un título general en la primera fila
        ws['A2'] = f"Cliente: {nombre_cliente}, Año: {año}"
        ws.merge_cells('A2:C2')  # Combina las celdas A1, B1, y C1
        ws['A2'].font = openpyxl.styles.Font(bold=True)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        ws.append([f"Fecha de Expedición: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.merge_cells('A3:C3')
        ws['A3'].font = openpyxl.styles.Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal="center")

        
                # Unir celdas desde fila 1
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)

        # Centrar texto en fila 1
        ws.cell(row=1, column=3).alignment = Alignment(horizontal="center") 

        # Ajustar ancho de columna del nombre  
        ws.column_dimensions['A'].auto_size = True
        # Estilo de celda para centrar texto
        align_center = Alignment(horizontal="center")
         # Agregar encabezados
        headers = ['Codigo', 'Descripcion', 'Monto']
        ws.append(headers)
        # Aplicar estilos a las celdas
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
            for cell in row:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = align_center

        # Escribir DataFrame
        for row in df.itertuples(index=False):
            ws.append(list(row))

        # Guardar el archivo Excel
        carpeta_destino = 'C:/Users/Naty/OneDrive/Documentos'
        nombre_archivo = f'balance_{nombre_cliente}_{año}_{fecha_hora_actual}.xlsx'
        ruta_archivo = os.path.join(carpeta_destino, nombre_archivo)
        wb.save(ruta_archivo)

        print("Directorio de trabajo actual:", os.getcwd())

        flash("Excel generado con éxito...")
        return send_file(ruta_archivo, as_attachment=True)
    except Exception as e:
        # En caso de error, puedes imprimir o manejar el error de alguna manera
        print(f"Error: {str(e)}")
        
            
@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.config.from_object(config['development'])
    csrf.init_app(app)
    app.register_error_handler(401, status_401)
    app.register_error_handler(404, status_404)
    

    app.run()